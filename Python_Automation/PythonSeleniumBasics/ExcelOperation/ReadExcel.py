import openpyxl

path = "C:\Pranay\PythonWorkspace\Python_Automation\Data1.xlsx"

# get workbook
workbook = openpyxl.load_workbook(path)

# Get sheet from workbook

sheet = workbook.active  # It will take active sheet from excel
# sheet = workbook.get_sheet_by_name("Sheet1")

# find how many number of rows and columns we have in sheet
rows = sheet.max_row
col = sheet.max_column

print(rows)
print(col)

# Get data from excel
for r in range(1, rows + 1):
    for c in range(1, col + 1):
        print(sheet.cell(row=r, column=c).value, end= "         ")
    print()
